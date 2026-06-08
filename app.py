# -*- coding: utf-8 -*-
"""
=============================================================================
  SIMANTAP UMKM - Sistem Manajemen Pantau UMKM
  Dinas Koperasi dan UKM Kabupaten Banyumas
  ---------------------------------------------------------------------------
  app.py : Dashboard pemantauan sentimen ulasan UMKM binaan (multi-halaman).

  Navigasi memakai st.session_state. Halaman:
    login -> overview -> daftar -> detail -> analisis -> laporan

  Prasyarat (jalankan lebih dulu):
    1. python train.py
    2. python generate_umkm.py
=============================================================================
"""

import json
import os
import re
from collections import Counter
from datetime import datetime
from io import BytesIO

import joblib
import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,  # noqa: F401  (diimpor sesuai spesifikasi)
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from Sastrawi.Stemmer.StemmerFactory import StemmerFactory

# ─────────────────────────────────────────────────────────────────────────────
# KONFIGURASI HALAMAN
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="SIMANTAP UMKM - Diskop Banyumas",
    page_icon="🏛️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
# KONSTANTA
# ─────────────────────────────────────────────────────────────────────────────
ADMIN_EMAIL = "admin@diskop.banyumaskab.go.id"
ADMIN_PASS  = "diskop2025"

# Palet warna design system
MERAH      = "#B91C1C"
HIJAU      = "#15803D"
TEKS       = "#111827"
TEKS2      = "#6B7280"
TEKS3      = "#9CA3AF"
BORDER     = "#E5E7EB"

# Gaya status UMKM: background, teks, border
STATUS_STYLE = {
    "Baik":            {"bg": "#DCFCE7", "fg": "#15803D", "bd": "#BBF7D0"},
    "Pantau":          {"bg": "#F3F4F6", "fg": "#6B7280", "bd": "#E5E7EB"},
    "Perlu Perhatian": {"bg": "#FEF3C7", "fg": "#B45309", "bd": "#FDE68A"},
    "Kritis":          {"bg": "#FEE2E2", "fg": "#B91C1C", "bd": "#FECACA"},
}

BULAN_ID = ["", "Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli",
            "Agustus", "September", "Oktober", "November", "Desember"]

MENU = [
    ("overview",      "Overview"),
    ("daftar",        "Daftar UMKM"),
    ("analisis",      "Uji Model"),
    ("perbandingan",  "Perbandingan Model"),
    ("laporan",       "Laporan"),
]

MODEL_FILES = {
    "Naive Bayes":   "nb_model.pkl",
    "SVM":           "svm_model.pkl",
    "Random Forest": "rf_model.pkl",
}
MODEL_DESC = {
    "Naive Bayes":   "Probabilistik, ringan dan cepat.",
    "SVM":           "SVM Linear terkalibrasi, akurasi tertinggi.",
    "Random Forest": "Ensemble 200 pohon keputusan, stabil.",
}

# Kamus keyword aspek (identik dengan generate_umkm.py)
ASPEK_KEYWORDS = {
    "Kualitas": [
        "kualitas", "bagus", "jelek", "rusak", "cacat", "sesuai", "tidak sesuai",
        "bahan", "material", "awet", "tahan", "rapuh", "original", "asli", "palsu",
        "mantap", "memuaskan", "kecewa", "mengecewakan", "sempurna", "buruk",
        "rasa", "tekstur", "warna", "bentuk", "ukuran", "fungsi", "berfungsi",
        "mati", "error", "lumayan", "oke", "produk", "barang",
    ],
    "Packaging": [
        "packing", "packaging", "kemasan", "bungkus", "kotak", "dos", "kardus",
        "bubble", "wrap", "aman", "selamat", "pecah", "retak", "penyok",
        "lecet", "rapi", "berantakan", "terbuka", "bocor",
    ],
    "Harga": [
        "harga", "mahal", "murah", "worth", "terjangkau", "overpriced",
        "ekonomis", "hemat", "bersaing", "sesuai harga", "nilai", "bayar",
        "ongkir", "gratis ongkir", "promo", "diskon", "biaya",
    ],
    "Pengiriman": [
        "pengiriman", "kirim", "ekspedisi", "kurir", "datang", "tiba",
        "sampai", "cepat", "lambat", "lama", "tepat waktu", "telat",
        "delay", "tracking", "resi", "j&t", "jne", "sicepat", "shopee express",
    ],
}


def deteksi_aspek(teks_asli: str) -> dict:
    """Deteksi aspek berbasis kata kunci pada teks ulasan asli."""
    teks = str(teks_asli).lower()
    return {aspek: any(kw in teks for kw in keywords)
            for aspek, keywords in ASPEK_KEYWORDS.items()}


# ─────────────────────────────────────────────────────────────────────────────
# CSS GLOBAL
# ─────────────────────────────────────────────────────────────────────────────
GLOBAL_CSS = """
<style>
/* Sembunyikan toolbar default Streamlit */
#MainMenu, footer, header { visibility: hidden; }
[data-testid="stToolbar"] { display: none; }

/* Background halaman */
.stApp { background-color: #F8F9FA; }

/* Sidebar — selalu tampil, override transform Streamlit */
section[data-testid="stSidebar"] {
    display: flex !important;
    transform: none !important;
    left: 0 !important;
    background-color: #FFFFFF !important;
    border-right: 1px solid #E5E7EB !important;
    width: 260px !important;
    min-width: 260px !important;
    visibility: visible !important;
    opacity: 1 !important;
}
section[data-testid="stSidebar"] * { color: #374151; }

/* Sembunyikan tombol collapse sidebar permanen */
[data-testid="stSidebarCollapseButton"],
[data-testid="stSidebarCollapsedControl"],
[data-testid="collapsedControl"],
button[data-testid="baseButton-headerNoPadding"] {
    display: none !important;
}

/* Padding kontainer */
.block-container { padding-top: 1.5rem !important; padding-bottom: 2rem !important; }

/* Tombol primer (aksi) */
.stButton > button, .stDownloadButton > button, .stFormSubmitButton > button {
    background-color: #B91C1C !important;
    color: #FFFFFF !important;
    border: none !important;
    border-radius: 6px !important;
    font-weight: 600 !important;
    padding: 0.5rem 1.25rem !important;
}
.stButton > button:hover, .stDownloadButton > button:hover,
.stFormSubmitButton > button:hover { background-color: #991B1B !important; }

/* Tombol full-width (pengganti width="stretch" yang tidak didukung di Streamlit <=1.40) */
.stButton > button { width: 100% !important; }
.stFormSubmitButton > button { width: 100% !important; }

/* Tombol navigasi sidebar */
section[data-testid="stSidebar"] .stButton > button {
    width: 100% !important;
    text-align: left !important;
    justify-content: flex-start !important;
    padding: 0.55rem 0.85rem !important;
    margin-bottom: 2px !important;
    box-shadow: none !important;
}
section[data-testid="stSidebar"] button[kind="primary"] {
    background-color: #FEE2E2 !important;
    color: #B91C1C !important;
    border-left: 3px solid #B91C1C !important;
    font-weight: 700 !important;
}
section[data-testid="stSidebar"] button[kind="secondary"] {
    background-color: transparent !important;
    color: #374151 !important;
    font-weight: 500 !important;
}
section[data-testid="stSidebar"] button[kind="secondary"]:hover {
    background-color: #F9FAFB !important;
    color: #B91C1C !important;
}

/* Input */
.stTextInput > div > div > input, .stTextArea textarea {
    border: 1px solid #D1D5DB !important;
    border-radius: 6px !important;
    background: #FFFFFF !important;
    color: #111827 !important;
}
.stSelectbox > div > div {
    border: 1px solid #D1D5DB !important;
    border-radius: 6px !important;
    background: #FFFFFF !important;
    color: #111827 !important;
}
[data-baseweb="select"] span,
[data-baseweb="select"] > div,
.stSelectbox [data-baseweb="select"] * {
    color: #111827 !important;
}
.stTextInput label, .stTextArea label, .stSelectbox label {
    color: #374151 !important; font-weight: 600 !important;
}

/* Tabel & dataframe */
.stDataFrame { border: 1px solid #E5E7EB; border-radius: 8px; }

/* Divider */
hr { border: none; border-top: 1px solid #E5E7EB; margin: 1rem 0; }

/* Tombol di luar sidebar — pastikan teks tidak wrap */
.block-container .stButton > button {
    white-space: nowrap !important;
    overflow: hidden !important;
    text-overflow: ellipsis !important;
    min-height: 2.2rem !important;
    line-height: 1.2 !important;
}

/* Alert (warning/info/error/success) — pastikan teks gelap & kontras
   pada background terang (mengatasi default theme yang tidak konsisten) */
.stAlert,
.stAlert p,
.stAlert span,
.stAlert div,
.stAlert strong,
.stAlert b,
.stAlert li {
    color: #1F2937 !important;
}
/* Warning khusus: latar kuning muda, teks coklat gelap untuk kontras tinggi */
.stAlert[data-baseweb="notification"][kind="warning"],
div[data-testid="stAlert"]:has([kind="warning"]) {
    background-color: #FFFBEB !important;
    border: 1px solid #FDE68A !important;
}
.stAlert[data-baseweb="notification"][kind="warning"] *,
div[data-testid="stAlert"]:has([kind="warning"]) * {
    color: #92400E !important;
}

/* Tabel HTML mentah di markdown — pastikan teks default cell gelap
   sehingga NAMA UMKM tidak tampil putih-di-putih. Hanya menarget tabel
   yang TIDAK memiliki class kustom (mis. .tbl-preproses), supaya tidak
   menimpa warna teks header putih pada tabel ber-class sendiri. */
.block-container table:not([class]) td,
.block-container table:not([class]) th { color: #111827; }

/* Expander header — pastikan label "Cara Kerja Preprocessing" jelas */
[data-testid="stExpander"] summary,
[data-testid="stExpander"] details > summary,
.streamlit-expanderHeader {
    color: #111827 !important;
    font-weight: 700 !important;
    background: #FFFFFF !important;
    border: 1px solid #E5E7EB !important;
    border-radius: 8px !important;
}
[data-testid="stExpander"] summary p,
[data-testid="stExpander"] summary span,
[data-testid="stExpander"] summary div {
    color: #111827 !important;
    font-weight: 700 !important;
}
[data-testid="stExpander"] details[open] > summary {
    border-bottom-left-radius: 0 !important;
    border-bottom-right-radius: 0 !important;
    border-bottom: 1px solid #E5E7EB !important;
}

/* Tabel "Cara Kerja Preprocessing" — semua warna dikunci !important
   agar tidak bertabrakan dengan default theme Streamlit */
.tbl-preproses {
    width: 100% !important;
    border-collapse: collapse !important;
    background: #FFFFFF !important;
    border: 1px solid #E5E7EB !important;
    border-radius: 8px !important;
    overflow: hidden !important;
    font-size: 0.9rem !important;
    margin-top: 0.4rem !important;
}
.tbl-preproses thead tr { background: #B91C1C !important; }
.tbl-preproses thead th {
    color: #FFFFFF !important;
    padding: 11px 14px !important;
    font-size: 0.78rem !important;
    font-weight: 700 !important;
    text-align: left !important;
    letter-spacing: 0.03em !important;
    border: none !important;
    text-transform: uppercase !important;
}
.tbl-preproses thead th.col-no { text-align: center !important; width: 56px !important; }
.tbl-preproses tbody td {
    padding: 11px 14px !important;
    border-bottom: 1px solid #E5E7EB !important;
    vertical-align: top !important;
    background: #FFFFFF !important;
}
.tbl-preproses tbody tr:nth-child(even) td { background: #F9FAFB !important; }
.tbl-preproses tbody tr:last-child td { border-bottom: none !important; }
.tbl-preproses tbody td.cell-no {
    color: #6B7280 !important;
    font-weight: 700 !important;
    text-align: center !important;
    width: 56px !important;
}
.tbl-preproses tbody td.cell-asli { color: #374151 !important; }
.tbl-preproses tbody td.cell-hasil {
    color: #166534 !important;
    font-weight: 700 !important;
    background: #F0FDF4 !important;
}
.tbl-preproses tbody tr:nth-child(even) td.cell-hasil {
    background: #DCFCE7 !important;
}

/* Catatan 4-tahap di bawah tabel */
.note-preproses {
    font-size: 0.85rem !important;
    color: #374151 !important;
    margin-top: 12px !important;
    padding: 10px 14px !important;
    background: #F9FAFB !important;
    border-left: 4px solid #B91C1C !important;
    border-radius: 4px !important;
}
.note-preproses b { color: #111827 !important; }
.note-preproses .tahap { color: #B91C1C !important; font-weight: 700 !important; }
</style>
"""
st.markdown(GLOBAL_CSS, unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────────────────────────────────────
if "halaman" not in st.session_state:
    st.session_state.halaman = "login"
if "login_ok" not in st.session_state:
    st.session_state.login_ok = False
if "umkm_dipilih" not in st.session_state:
    st.session_state.umkm_dipilih = None
if "daftar_page" not in st.session_state:
    st.session_state.daftar_page = 0

# ─────────────────────────────────────────────────────────────────────────────
# LOADER RESOURCE (cache)
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_resource(show_spinner=False)
def load_stemmer():
    return StemmerFactory().create_stemmer()


@st.cache_resource(show_spinner=False)
def load_stopwords():
    if os.path.exists("stopwords_set.pkl"):
        return joblib.load("stopwords_set.pkl")
    return set()


@st.cache_resource(show_spinner="Memuat model …")
def load_models():
    res = {"vectorizer": joblib.load("vectorizer.pkl")}
    for nama, berkas in MODEL_FILES.items():
        res[nama] = joblib.load(berkas)
    return res


@st.cache_data(show_spinner=False)
def load_umkm_data():
    with open("umkm_data.json", encoding="utf-8") as f:
        return json.load(f)


@st.cache_data(show_spinner=False)
def load_evaluation():
    if os.path.exists("evaluation_results.json"):
        with open("evaluation_results.json", encoding="utf-8") as f:
            return json.load(f)
    return {}


def preprocess(text: str) -> str:
    """Pipeline preprocessing — IDENTIK dengan train.py."""
    stopwords = load_stopwords()
    stemmer = load_stemmer()
    text = str(text).lower()
    text = re.sub(r"http\S+|www\.\S+", " ", text)
    text = re.sub(r"@\w+|#\w+", " ", text)
    text = re.sub(r"\d+", " ", text)
    text = re.sub(r"[^\w\s]", " ", text)
    text = text.replace("_", " ")
    tokens = text.split()
    tokens = [t for t in tokens if t not in stopwords and len(t) > 1]
    tokens = [stemmer.stem(t) for t in tokens]
    return " ".join(tokens)


def butuh_umkm_data():
    """Kembalikan data UMKM; bila belum ada, beri instruksi & hentikan halaman."""
    if not os.path.exists("umkm_data.json"):
        st.info(
            "Data UMKM belum tersedia. Jalankan terlebih dahulu di terminal:\n\n"
            "`python train.py`  lalu  `python generate_umkm.py`"
        )
        st.stop()
    return load_umkm_data()


# ─────────────────────────────────────────────────────────────────────────────
# KOMPONEN HTML
# ─────────────────────────────────────────────────────────────────────────────
def judul_halaman(teks, sub=""):
    html = f'<div style="font-size:1.5rem;font-weight:700;color:{TEKS};">{teks}</div>'
    if sub:
        html += f'<div style="color:{TEKS2};font-size:0.9rem;margin-top:2px;">{sub}</div>'
    st.markdown(html, unsafe_allow_html=True)


def judul_seksi(teks):
    st.markdown(
        f'<div style="font-size:1.05rem;font-weight:700;color:{TEKS};'
        f'margin:0.6rem 0 0.4rem;">{teks}</div>',
        unsafe_allow_html=True,
    )


def badge_status(status):
    s = STATUS_STYLE[status]
    return (f'<span style="background:{s["bg"]};color:{s["fg"]};'
            f'border:1px solid {s["bd"]};padding:3px 11px;border-radius:12px;'
            f'font-size:0.78rem;font-weight:600;white-space:nowrap;">{status}</span>')


def metric_card(label, value, warna):
    return (f'<div style="background:#FFFFFF;border:1px solid {BORDER};'
            f'border-radius:8px;padding:1rem 1.25rem;'
            f'box-shadow:0 1px 3px rgba(0,0,0,0.06);">'
            f'<div style="font-size:2rem;font-weight:800;color:{warna};'
            f'line-height:1;">{value}</div>'
            f'<div style="font-size:0.8rem;color:{TEKS2};margin-top:6px;'
            f'font-weight:600;">{label}</div></div>')


def skor_bar(skor, status):
    fg = STATUS_STYLE[status]["fg"]
    return (f'<div style="display:flex;align-items:center;gap:8px;">'
            f'<div style="flex:1;background:#F3F4F6;border-radius:6px;'
            f'height:8px;overflow:hidden;">'
            f'<div style="width:{skor}%;height:100%;background:{fg};"></div></div>'
            f'<span style="font-weight:700;color:{TEKS};font-size:0.85rem;">'
            f'{skor}</span></div>')


def chip(teks, jenis):
    bg, fg = ("#DCFCE7", "#15803D") if jenis == "pos" else ("#FEE2E2", "#B91C1C")
    return (f'<span style="display:inline-block;background:{bg};color:{fg};'
            f'padding:3px 10px;border-radius:12px;font-size:0.78rem;'
            f'margin:2px 3px;font-weight:500;">{teks}</span>')


# ─────────────────────────────────────────────────────────────────────────────
# DETEKSI POLA KELUHAN (untuk alert)
# ─────────────────────────────────────────────────────────────────────────────
def cek_pola_keluhan(daftar_umkm):
    """Cari aspek yang menjadi masalah utama pada >50% UMKM bermasalah."""
    bermasalah = [u for u in daftar_umkm
                  if u["status"] in ("Kritis", "Perlu Perhatian")]
    if not bermasalah:
        return None
    hitung = Counter(u["masalah_utama"] for u in bermasalah
                     if u["masalah_utama"] != "Tidak ada")
    if not hitung:
        return None
    aspek, jumlah = hitung.most_common(1)[0]
    if jumlah / len(bermasalah) > 0.5:
        return {"aspek": aspek, "jumlah": jumlah, "total": len(bermasalah)}
    return None


SARAN_WORKSHOP = {
    "Packaging":  "pertimbangkan workshop kemasan bersama",
    "Kualitas":   "pertimbangkan pendampingan quality control terpadu",
    "Harga":      "pertimbangkan pelatihan manajemen biaya bersama",
    "Pengiriman": "pertimbangkan kerja sama mitra ekspedisi kolektif",
}


# =============================================================================
# HALAMAN 0 — LOGIN
# =============================================================================
def render_login():
    # Sembunyikan sidebar di halaman login — override transform DAN display
    st.markdown("""
    <style>
    section[data-testid="stSidebar"] {
        display: none !important;
        transform: translateX(-100%) !important;
        visibility: hidden !important;
    }
    [data-testid="stSidebarCollapseButton"],
    [data-testid="stSidebarCollapsedControl"],
    [data-testid="collapsedControl"] { display: none !important; }
    </style>
    """, unsafe_allow_html=True)
    c1, c2, c3 = st.columns([1, 1.3, 1])
    with c2:
        st.markdown(
            f"""
            <div style="text-align:center;margin-top:3rem;margin-bottom:0.5rem;">
                <div style="font-size:1.8rem;font-weight:800;color:{MERAH};
                    letter-spacing:0.5px;">SIMANTAP UMKM</div>
                <div style="font-size:1rem;font-weight:600;color:{TEKS};
                    margin-top:8px;">Dinas Koperasi dan UKM Kabupaten Banyumas</div>
                <div style="font-size:0.85rem;color:{TEKS2};margin-top:4px;">
                    Sistem Pemantauan Sentimen Ulasan UMKM Binaan</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        with st.container(border=True):
            st.markdown(
                f'<div style="font-weight:700;color:{TEKS};font-size:1.05rem;'
                f'margin-bottom:0.3rem;">Masuk ke Dashboard</div>',
                unsafe_allow_html=True,
            )
            with st.form("form_login"):
                email = st.text_input("Email", placeholder="nama@diskop.banyumaskab.go.id")
                sandi = st.text_input("Password", type="password",
                                      placeholder="Masukkan kata sandi")
                masuk = st.form_submit_button("Masuk")
            if masuk:
                if email.strip() == ADMIN_EMAIL and sandi == ADMIN_PASS:
                    st.session_state.login_ok = True
                    st.session_state.halaman = "overview"
                    st.rerun()
                else:
                    st.error("Email atau password salah.")
        st.markdown(
            f'<div style="text-align:center;color:{TEKS3};font-size:0.78rem;'
            f'margin-top:0.8rem;">Demo: gunakan email dan password yang tertera '
            f'di dokumen teknis.</div>',
            unsafe_allow_html=True,
        )


# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR NAVIGASI
# ─────────────────────────────────────────────────────────────────────────────
def render_sidebar():
    # GLOBAL_CSS sudah mengurus sidebar visibility via transform override.
    # Tidak perlu injeksi CSS tambahan di sini.
    with st.sidebar:
        st.markdown(
            f"""
            <div style="padding:0.4rem 0 0.8rem;">
                <div style="font-size:1.15rem;font-weight:800;color:{MERAH};">
                    SIMANTAP UMKM</div>
                <div style="font-size:0.78rem;color:{TEKS2};">Diskop Banyumas</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        st.markdown("<hr style='margin:0.3rem 0 0.7rem;'>", unsafe_allow_html=True)

        # Halaman 'detail' dianggap bagian dari menu 'Daftar UMKM'
        aktif = "daftar" if st.session_state.halaman == "detail" else st.session_state.halaman
        for kode, label in MENU:
            tipe = "primary" if kode == aktif else "secondary"
            if st.button(label, key=f"nav_{kode}", type=tipe):
                st.session_state.halaman = kode
                if kode == "daftar":
                    st.session_state.daftar_page = 0
                st.rerun()

        st.markdown("<hr style='margin:0.7rem 0;'>", unsafe_allow_html=True)

        diperbarui = "hari ini"
        if os.path.exists("umkm_data.json"):
            try:
                diperbarui = load_umkm_data().get("generated_at", "hari ini")
            except Exception:
                pass
        st.markdown(
            f'<div style="font-size:0.74rem;color:{TEKS3};margin-bottom:0.6rem;">'
            f'Diperbarui: {diperbarui}</div>',
            unsafe_allow_html=True,
        )
        if st.button("Keluar", key="btn_keluar", type="secondary"):
            st.session_state.login_ok = False
            st.session_state.halaman = "login"
            st.rerun()


# =============================================================================
# HALAMAN 1 — OVERVIEW
# =============================================================================
def page_overview():
    data = butuh_umkm_data()
    umkm = data["umkm"]
    ring = data["ringkasan"]

    now = datetime.now()
    judul_halaman(
        "Ringkasan Pemantauan UMKM",
        f"Periode: {BULAN_ID[now.month]} {now.year} · "
        f"{data['total_umkm']} UMKM terpantau",
    )
    st.markdown("<div style='height:0.6rem;'></div>", unsafe_allow_html=True)

    # ── 4 metric card ──
    cols = st.columns(4)
    metrik = [
        ("Kritis", ring["kritis"], MERAH),
        ("Perlu Perhatian", ring["perlu_perhatian"], "#B45309"),
        ("Pantau", ring["pantau"], TEKS2),
        ("Baik", ring["baik"], HIJAU),
    ]
    for col, (lbl, val, warna) in zip(cols, metrik):
        col.markdown(metric_card(lbl, val, warna), unsafe_allow_html=True)

    st.markdown("<div style='margin-top: 1.5rem;'></div>", unsafe_allow_html=True)

    # ── Alert pola keluhan ──
    pola = cek_pola_keluhan(umkm)
    if pola:
        saran = SARAN_WORKSHOP.get(pola["aspek"], "pertimbangkan program bersama")
        st.markdown("<div style='height:0.6rem;'></div>", unsafe_allow_html=True)
        st.warning(
            f"{pola['jumlah']} dari {pola['total']} UMKM bermasalah memiliki "
            f"masalah utama di aspek **{pola['aspek']}** — {saran}."
        )

    st.markdown("<hr>", unsafe_allow_html=True)

    # ── Grafik aspek keluhan ──
    judul_seksi("Aspek Dominan Keluhan UMKM Bermasalah")
    bermasalah = [u for u in umkm if u["status"] in ("Kritis", "Perlu Perhatian")]
    if bermasalah:
        persen = {}
        for aspek in ASPEK_KEYWORDS:
            jml = sum(1 for u in bermasalah
                      if u["aspek"][aspek]["keluhan_count"] > 0)
            persen[aspek] = round(jml / len(bermasalah) * 100, 1)
        urut = sorted(persen.items(), key=lambda x: x[1])
        nama = [k for k, _ in urut]
        nilai = [v for _, v in urut]
        warna = ["#B91C1C" if v > 50 else ("#15803D" if v < 30 else "#D97706")
                 for v in nilai]
        fig = go.Figure(go.Bar(
            x=nilai, y=nama, orientation="h", marker_color=warna,
            text=[f"{v}%" for v in nilai], textposition="outside",
        ))
        fig.update_layout(
            template="plotly_white", height=280,
            paper_bgcolor="white",
            plot_bgcolor="white",
            font=dict(color="#374151", family="sans-serif"),
            margin=dict(l=16, r=16, t=32, b=16),
            xaxis=dict(title="Persentase UMKM Bermasalah", range=[0, 110]),
            yaxis=dict(title=""),
        )
        fig.update_xaxes(
            gridcolor="#F3F4F6",
            linecolor="#E5E7EB",
            tickfont=dict(color="#6B7280"),
        )
        fig.update_yaxes(
            gridcolor="#F3F4F6",
            linecolor="#E5E7EB",
            tickfont=dict(color="#374151"),
        )
        st.plotly_chart(fig, use_container_width=True)
        st.caption("Persentase UMKM berstatus Kritis / Perlu Perhatian yang "
                   "memiliki keluhan pada tiap aspek.")
    else:
        st.info("Tidak ada UMKM berstatus Kritis atau Perlu Perhatian.")

    st.markdown("<hr>", unsafe_allow_html=True)

    st.markdown("<div style='margin-top: 1rem;'></div>", unsafe_allow_html=True)

    # ── Tabel UMKM kritis ──
    judul_seksi("UMKM Berstatus Kritis")
    kritis = sorted([u for u in umkm if u["status"] == "Kritis"],
                    key=lambda u: u["skor"])[:5]
    if kritis:
        baris = ""
        for u in kritis:
            baris += (
                f'<tr>'
                f'<td style="padding:9px 12px;border-bottom:1px solid {BORDER};'
                f'color:{TEKS};font-weight:600;">'
                f'{u["nama"]}</td>'
                f'<td style="padding:9px 12px;border-bottom:1px solid {BORDER};'
                f'color:{TEKS2};">{u["kategori"]}</td>'
                f'<td style="padding:9px 12px;border-bottom:1px solid {BORDER};'
                f'font-weight:700;color:{MERAH};">{u["skor"]}</td>'
                f'<td style="padding:9px 12px;border-bottom:1px solid {BORDER};'
                f'color:{TEKS2};">{u["masalah_utama"]}</td>'
                f'<td style="padding:9px 12px;border-bottom:1px solid {BORDER};">'
                f'{badge_status(u["status"])}</td>'
                f'</tr>'
            )
        st.markdown(
            f"""
            <table style="width:100%;border-collapse:collapse;background:#FFFFFF;
                border:1px solid {BORDER};border-radius:8px;overflow:hidden;
                font-size:0.88rem;">
                <tr style="background:#F9FAFB;">
                    <th style="padding:9px 12px;text-align:left;color:{TEKS2};
                        font-size:0.78rem;">NAMA UMKM</th>
                    <th style="padding:9px 12px;text-align:left;color:{TEKS2};
                        font-size:0.78rem;">KATEGORI</th>
                    <th style="padding:9px 12px;text-align:left;color:{TEKS2};
                        font-size:0.78rem;">SKOR</th>
                    <th style="padding:9px 12px;text-align:left;color:{TEKS2};
                        font-size:0.78rem;">MASALAH UTAMA</th>
                    <th style="padding:9px 12px;text-align:left;color:{TEKS2};
                        font-size:0.78rem;">STATUS</th>
                </tr>
                {baris}
            </table>
            """,
            unsafe_allow_html=True,
        )
    else:
        st.success("Tidak ada UMKM berstatus Kritis saat ini.")

    st.markdown("<div style='height:0.8rem;'></div>", unsafe_allow_html=True)
    cbtn = st.columns([1, 1, 1])
    with cbtn[0]:
        if st.button("Lihat semua UMKM", key="ov_lihat"):
            st.session_state.halaman = "daftar"
            st.session_state.daftar_page = 0
            st.rerun()


# =============================================================================
# HALAMAN 2 — DAFTAR UMKM
# =============================================================================
def page_daftar():
    data = butuh_umkm_data()
    umkm = data["umkm"]

    judul_halaman("Daftar UMKM Binaan",
                  f"Total {len(umkm)} UMKM dalam pemantauan")
    st.markdown("<div style='height:0.6rem;'></div>", unsafe_allow_html=True)

    # ── Filter ──
    f1, f2, f3 = st.columns(3)
    with f1:
        f_status = st.selectbox(
            "Status",
            ["Semua", "Kritis", "Perlu Perhatian", "Pantau", "Baik"],
        )
    with f2:
        kategori_list = ["Semua"] + sorted({u["kategori"] for u in umkm})
        f_kategori = st.selectbox("Kategori", kategori_list)
    with f3:
        f_urut = st.selectbox(
            "Urutkan", ["Skor Terendah", "Skor Tertinggi", "Nama A-Z"]
        )

    # ── Terapkan filter ──
    hasil = list(umkm)
    if f_status != "Semua":
        hasil = [u for u in hasil if u["status"] == f_status]
    if f_kategori != "Semua":
        hasil = [u for u in hasil if u["kategori"] == f_kategori]
    if f_urut == "Skor Terendah":
        hasil.sort(key=lambda u: u["skor"])
    elif f_urut == "Skor Tertinggi":
        hasil.sort(key=lambda u: u["skor"], reverse=True)
    else:
        hasil.sort(key=lambda u: u["nama"])

    # ── Alert pola keluhan (untuk hasil filter) ──
    pola = cek_pola_keluhan(hasil)
    if pola and f_status in ("Semua", "Kritis", "Perlu Perhatian"):
        saran = SARAN_WORKSHOP.get(pola["aspek"], "pertimbangkan program bersama")
        st.warning(
            f"{pola['jumlah']} dari {pola['total']} UMKM bermasalah pada hasil "
            f"ini memiliki masalah utama di aspek **{pola['aspek']}** — {saran}."
        )

    st.caption(f"Menampilkan {len(hasil)} UMKM.")

    if not hasil:
        st.info("Tidak ada UMKM yang cocok dengan filter.")
        return

    # ── Pagination ──
    per_halaman = 20
    total_halaman = (len(hasil) - 1) // per_halaman + 1
    if st.session_state.daftar_page >= total_halaman:
        st.session_state.daftar_page = 0
    page = st.session_state.daftar_page
    potongan = hasil[page * per_halaman:(page + 1) * per_halaman]

    # ── Header tabel ──
    lebar = [0.5, 2.6, 1.3, 1.5, 1.9, 1.4, 1.3, 1.4]
    head = st.columns(lebar)
    for col, teks in zip(
        head, ["No", "Nama UMKM", "Kategori", "Lokasi", "Skor",
               "Masalah", "Status", ""]
    ):
        col.markdown(
            f'<div style="font-size:0.74rem;font-weight:700;color:{TEKS2};'
            f'text-transform:uppercase;padding:4px 0;">{teks}</div>',
            unsafe_allow_html=True,
        )
    st.markdown(f"<hr style='margin:0.1rem 0 0.3rem;'>", unsafe_allow_html=True)

    # ── Baris tabel ──
    for i, u in enumerate(potongan):
        nomor = page * per_halaman + i + 1
        row = st.columns(lebar)
        row[0].markdown(
            f'<div style="padding:6px 0;color:{TEKS2};">{nomor}</div>',
            unsafe_allow_html=True)
        row[1].markdown(
            f'<div style="padding:6px 0;font-weight:600;color:{TEKS};">'
            f'{u["nama"]}</div>'
            f'<div style="font-size:0.74rem;color:{TEKS3};">{u["pemilik"]}</div>',
            unsafe_allow_html=True)
        row[2].markdown(
            f'<div style="padding:6px 0;color:{TEKS2};font-size:0.85rem;">'
            f'{u["kategori"]}</div>', unsafe_allow_html=True)
        row[3].markdown(
            f'<div style="padding:6px 0;color:{TEKS2};font-size:0.85rem;">'
            f'{u["lokasi"]}</div>', unsafe_allow_html=True)
        row[4].markdown(
            f'<div style="padding:8px 0;">{skor_bar(u["skor"], u["status"])}</div>',
            unsafe_allow_html=True)
        row[5].markdown(
            f'<div style="padding:6px 0;color:{TEKS2};font-size:0.85rem;">'
            f'{u["masalah_utama"]}</div>', unsafe_allow_html=True)
        row[6].markdown(
            f'<div style="padding:6px 0;">{badge_status(u["status"])}</div>',
            unsafe_allow_html=True)
        if row[7].button("Detail", key=f"detail_{u['id']}"):
            st.session_state.umkm_dipilih = u["id"]
            st.session_state.halaman = "detail"
            st.rerun()
        st.markdown(
            f"<hr style='margin:0;border-top:1px solid #F3F4F6;'>",
            unsafe_allow_html=True)

    # ── Kontrol pagination ──
    if total_halaman > 1:
        st.markdown("<div style='height:0.5rem;'></div>", unsafe_allow_html=True)
        pg = st.columns([1, 1, 3, 1])
        with pg[0]:
            if st.button("◀ Sebelumnya", key="pg_prev", disabled=(page == 0)):
                st.session_state.daftar_page -= 1
                st.rerun()
        with pg[1]:
            if st.button("Berikutnya ▶", key="pg_next",
                         disabled=(page >= total_halaman - 1)):
                st.session_state.daftar_page += 1
                st.rerun()
        with pg[2]:
            st.markdown(
                f'<div style="padding-top:8px;color:{TEKS2};font-size:0.85rem;">'
                f'Halaman {page + 1} dari {total_halaman}</div>',
                unsafe_allow_html=True)


# =============================================================================
# HALAMAN 3 — DETAIL UMKM
# =============================================================================
def page_detail():
    data = butuh_umkm_data()
    umkm_id = st.session_state.umkm_dipilih
    umkm = next((u for u in data["umkm"] if u["id"] == umkm_id), None)

    if umkm is None:
        st.warning("UMKM tidak ditemukan.")
        if st.button("← Kembali ke Daftar"):
            st.session_state.halaman = "daftar"
            st.rerun()
        return

    # ── Tombol kembali ──
    if st.button("← Kembali ke Daftar UMKM", key="dt_kembali"):
        st.session_state.halaman = "daftar"
        st.rerun()

    # ── Header ──
    s = STATUS_STYLE[umkm["status"]]
    hcol = st.columns([3, 1])
    with hcol[0]:
        st.markdown(
            f'<div style="font-size:1.5rem;font-weight:800;color:{TEKS};">'
            f'{umkm["nama"]}</div>'
            f'<div style="color:{TEKS2};font-size:0.9rem;margin-top:3px;">'
            f'{umkm["kategori"]} · {umkm["lokasi"]}</div>'
            f'<div style="color:{TEKS2};font-size:0.9rem;">'
            f'Produk: {umkm["produk_utama"]}</div>',
            unsafe_allow_html=True)
    with hcol[1]:
        st.markdown(
            f'<div style="text-align:right;">'
            f'<div style="font-size:2.2rem;font-weight:800;color:{s["fg"]};">'
            f'{umkm["skor"]}</div>'
            f'<div style="margin-top:2px;">{badge_status(umkm["status"])}</div>'
            f'</div>',
            unsafe_allow_html=True)

    st.markdown("<hr>", unsafe_allow_html=True)

    kiri, kanan = st.columns([3, 2], gap="large")

    # ── KIRI: aspek & contoh ulasan ──
    with kiri:
        judul_seksi("Sentimen per Aspek")
        for aspek in ["Kualitas", "Packaging", "Harga", "Pengiriman"]:
            d = umkm["aspek"][aspek]
            if not d["terdeteksi"]:
                st.markdown(
                    f'<div style="margin-bottom:14px;">'
                    f'<div style="font-weight:600;color:{TEKS};">{aspek}</div>'
                    f'<div style="color:{TEKS3};font-size:0.82rem;'
                    f'font-style:italic;">Data belum cukup — aspek tidak '
                    f'disebut dalam ulasan</div></div>',
                    unsafe_allow_html=True)
            else:
                sp = d["skor_positif"]
                warna = HIJAU if sp > 60 else ("#D97706" if sp >= 40 else MERAH)
                st.markdown(
                    f'<div style="margin-bottom:14px;">'
                    f'<div style="display:flex;justify-content:space-between;">'
                    f'<span style="font-weight:600;color:{TEKS};">{aspek}</span>'
                    f'<span style="font-weight:700;color:{warna};">{sp}% '
                    f'positif</span></div>'
                    f'<div style="background:#F3F4F6;border-radius:6px;height:10px;'
                    f'overflow:hidden;margin-top:4px;">'
                    f'<div style="width:{sp}%;height:100%;background:{warna};">'
                    f'</div></div>'
                    f'<div style="font-size:0.76rem;color:{TEKS3};margin-top:3px;">'
                    f'{d.get("positif_count", 0)} positif · {d["keluhan_count"]} negatif · dari {d.get("total_mention", 0)} ulasan yang menyebut aspek ini</div>'
                    f'</div>',
                    unsafe_allow_html=True)

        st.markdown("<div style='height:0.4rem;'></div>", unsafe_allow_html=True)
        judul_seksi("Ulasan Terpilih")
        if umkm["contoh_review_positif"]:
            st.markdown(
                f'<div style="background:#FFFFFF;border:1px solid {BORDER};'
                f'border-left:3px solid {HIJAU};border-radius:6px;padding:10px 14px;'
                f'margin-bottom:10px;">'
                f'<span style="background:#DCFCE7;color:{HIJAU};padding:2px 8px;'
                f'border-radius:10px;font-size:0.72rem;font-weight:600;">'
                f'Positif</span>'
                f'<div style="color:{TEKS};font-size:0.86rem;margin-top:6px;'
                f'line-height:1.5;">{umkm["contoh_review_positif"]}</div></div>',
                unsafe_allow_html=True)
        if umkm["contoh_review_negatif"]:
            st.markdown(
                f'<div style="background:#FFFFFF;border:1px solid {BORDER};'
                f'border-left:3px solid {MERAH};border-radius:6px;padding:10px 14px;">'
                f'<span style="background:#FEE2E2;color:{MERAH};padding:2px 8px;'
                f'border-radius:10px;font-size:0.72rem;font-weight:600;">'
                f'Negatif</span>'
                f'<div style="color:{TEKS};font-size:0.86rem;margin-top:6px;'
                f'line-height:1.5;">{umkm["contoh_review_negatif"]}</div></div>',
                unsafe_allow_html=True)
        if not umkm["contoh_review_negatif"]:
            st.caption("Tidak ada ulasan negatif pada sampel UMKM ini.")

    # ── KANAN: kata kunci, rekomendasi, statistik ──
    with kanan:
        judul_seksi("Kata Kunci Dominan")
        chips = ""
        for k in umkm["kata_kunci_positif"]:
            chips += chip(k, "pos")
        for k in umkm["kata_kunci_negatif"]:
            chips += chip(k, "neg")
        st.markdown(f'<div>{chips if chips else "—"}</div>',
                    unsafe_allow_html=True)

        st.markdown("<div style='height:0.7rem;'></div>", unsafe_allow_html=True)
        judul_seksi("Rekomendasi Tindakan Diskop")
        item = "".join(
            f'<li style="margin-bottom:5px;color:{TEKS};font-size:0.87rem;">{r}</li>'
            for r in umkm["rekomendasi"])
        st.markdown(
            f'<div style="border-left:3px solid {MERAH};padding:4px 0 4px 14px;">'
            f'<ol style="margin:0;padding-left:18px;">{item}</ol></div>',
            unsafe_allow_html=True)

        st.markdown("<div style='height:0.7rem;'></div>", unsafe_allow_html=True)
        judul_seksi("Statistik Ringkas")
        total = umkm["total_review"]
        pp = round(umkm["review_positif"] / total * 100) if total else 0
        pn = round(umkm["review_negatif"] / total * 100) if total else 0
        st.markdown(
            f'<div style="background:#FFFFFF;border:1px solid {BORDER};'
            f'border-radius:8px;padding:12px 16px;font-size:0.88rem;color:{TEKS};">'
            f'<div style="margin-bottom:4px;">Total ulasan dianalisis: '
            f'<b>{total}</b></div>'
            f'<div style="margin-bottom:4px;color:{HIJAU};">Positif: '
            f'<b>{umkm["review_positif"]}</b> ({pp}%)</div>'
            f'<div style="color:{MERAH};">Negatif: '
            f'<b>{umkm["review_negatif"]}</b> ({pn}%)</div></div>',
            unsafe_allow_html=True)


# =============================================================================
# HALAMAN 4 — ANALISIS ULASAN
# =============================================================================
def page_analisis():
    judul_halaman("Uji Model Klasifikasi",
                  "Verifikasi akurasi model sebelum data batch UMKM diproses")
    st.markdown("<div style='height:0.6rem;'></div>", unsafe_allow_html=True)

    st.markdown(
        f"""<div style="background:#EFF6FF;border:1px solid #BFDBFE;border-radius:8px;
            padding:12px 16px;margin-bottom:1.2rem;font-size:0.875rem;color:#1E40AF;">
            Gunakan halaman ini untuk menguji respons model secara langsung terhadap
            ulasan baru. Berguna untuk memverifikasi bahwa pipeline NLP berjalan
            benar sebelum data ulasan UMKM diproses secara batch.
        </div>""",
        unsafe_allow_html=True,
    )

    with st.expander("Cara Kerja Preprocessing"):
        contoh_preproses = [
            ("Produk BAGUS banget!! Pengiriman cepet, packing oke",
             "produk bagus kirim cepat packing oke"),
            ("Kecewa bgt.. barang ga sesuai deskripsi, harga mahal jg",
             "kecewa barang sesuai deskripsi harga mahal"),
            ("Udh beli 3x, kualitasnya konsisten & seller responsif banget",
             "beli kualitas konsisten seller responsif"),
            ("Pengiriman lama bgt sampe 2 minggu, tp barangnya oke lah",
             "kirim lama minggu barang oke"),
        ]
        baris_contoh = "".join(
            f'<tr>'
            f'<td class="cell-no">{i}</td>'
            f'<td class="cell-asli">{asli}</td>'
            f'<td class="cell-hasil">{hasil}</td>'
            f'</tr>'
            for i, (asli, hasil) in enumerate(contoh_preproses, start=1)
        )
        st.markdown(
            f"""
            <table class="tbl-preproses">
                <thead>
                    <tr>
                        <th class="col-no">NO</th>
                        <th>TEKS ASLI</th>
                        <th>HASIL PREPROCESSING</th>
                    </tr>
                </thead>
                <tbody>
                    {baris_contoh}
                </tbody>
            </table>
            """,
            unsafe_allow_html=True,
        )
        st.markdown(
            '<div class="note-preproses">'
            '<b>4 Tahap Preprocessing:</b> '
            '<span class="tahap">Lowercase</span> → '
            '<span class="tahap">Cleaning</span> '
            '(buang angka, tanda baca, karakter khusus) → '
            '<span class="tahap">Stopword Removal</span> '
            '(buang kata umum seperti "yang", "di", "ke") → '
            '<span class="tahap">Stemming</span> dengan '
            '<b>Sastrawi</b> (kata dasar bahasa Indonesia).'
            '</div>',
            unsafe_allow_html=True,
        )

    # Cek ketersediaan model
    kurang = [f for f in ["vectorizer.pkl"] + list(MODEL_FILES.values())
              if not os.path.exists(f)]
    if kurang:
        st.info(
            "Model belum tersedia. Jalankan terlebih dahulu di terminal:\n\n"
            "`python train.py`"
        )
        st.stop()

    kiri, kanan = st.columns([2, 3], gap="large")

    with kiri:
        model_pil = st.selectbox("Model klasifikasi",
                                 ["Naive Bayes", "SVM", "Random Forest"], index=1)
        st.caption(MODEL_DESC[model_pil])
        teks = st.text_area(
            "Masukkan ulasan untuk diuji", height=170,
            placeholder='Contoh: "Barang bagus, packing rapi, '
                        'pengiriman cepat. Sangat puas!"',
        )
        analisa = st.button("Analisis", key="btn_analisis")

    # ── Proses analisis ──
    hasil = None
    if analisa and teks.strip():
        with st.spinner("Memproses ulasan …"):
            res = load_models()
            vectorizer = res["vectorizer"]
            model = res[model_pil]
            diproses = preprocess(teks)
            vec = vectorizer.transform([diproses])
            kelas = list(model.classes_)
            proba = model.predict_proba(vec)[0]
            p_pos = float(proba[kelas.index("Positive")])
            p_neg = float(proba[kelas.index("Negative")])
            label = "Positive" if p_pos >= p_neg else "Negative"
            hasil = {
                "label": label, "p_pos": p_pos, "p_neg": p_neg,
                "diproses": diproses, "vec": vec, "vectorizer": vectorizer,
                "aspek": deteksi_aspek(teks),
            }

    with kanan:
        if hasil is None and analisa:
            st.warning("Masukkan teks ulasan terlebih dahulu.")
        elif hasil is None:
            st.markdown(
                f'<div style="background:#FFFFFF;border:1px dashed {BORDER};'
                f'border-radius:8px;padding:2rem;text-align:center;color:{TEKS3};">'
                f'Hasil analisis akan ditampilkan di sini setelah Anda '
                f'menekan tombol Analisis.</div>',
                unsafe_allow_html=True)
        else:
            menang = hasil["label"]
            conf = hasil["p_pos"] if menang == "Positive" else hasil["p_neg"]
            if menang == "Positive":
                bg, fg, teks_label = "#DCFCE7", HIJAU, "SENTIMEN POSITIF"
            else:
                bg, fg, teks_label = "#FEE2E2", MERAH, "SENTIMEN NEGATIF"
            st.markdown(
                f'<div style="background:{bg};border:1px solid {fg}33;'
                f'border-radius:8px;padding:1.1rem;text-align:center;">'
                f'<div style="font-size:1.3rem;font-weight:800;color:{fg};">'
                f'{teks_label}</div></div>',
                unsafe_allow_html=True)

            # Confidence bar
            pos_pct = hasil["p_pos"] * 100
            neg_pct = hasil["p_neg"] * 100
            st.markdown("<div style='height:0.6rem;'></div>",
                        unsafe_allow_html=True)
            st.markdown(
                f'<div style="font-weight:600;color:{TEKS};font-size:0.88rem;">'
                f'Keyakinan model: {conf * 100:.1f}%</div>'
                f'<div style="display:flex;height:22px;border-radius:6px;'
                f'overflow:hidden;border:1px solid {BORDER};margin-top:5px;">'
                f'<div style="width:{pos_pct:.1f}%;background:{HIJAU};color:#FFF;'
                f'font-size:0.72rem;font-weight:700;display:flex;'
                f'align-items:center;justify-content:center;">'
                f'{pos_pct:.0f}%</div>'
                f'<div style="width:{neg_pct:.1f}%;background:{MERAH};color:#FFF;'
                f'font-size:0.72rem;font-weight:700;display:flex;'
                f'align-items:center;justify-content:center;">'
                f'{neg_pct:.0f}%</div></div>'
                f'<div style="display:flex;justify-content:space-between;'
                f'font-size:0.74rem;color:{TEKS2};margin-top:3px;">'
                f'<span>Positif</span><span>Negatif</span></div>',
                unsafe_allow_html=True)

            # Aspek terdeteksi
            st.markdown("<div style='height:0.6rem;'></div>",
                        unsafe_allow_html=True)
            st.markdown(
                f'<div style="font-weight:600;color:{TEKS};font-size:0.88rem;'
                f'margin-bottom:4px;">Aspek terdeteksi</div>',
                unsafe_allow_html=True)
            terdeteksi = [a for a, on in hasil["aspek"].items() if on]
            if terdeteksi:
                chips = "".join(
                    f'<span style="display:inline-block;background:#FEE2E2;'
                    f'color:{MERAH};padding:3px 11px;border-radius:12px;'
                    f'font-size:0.78rem;margin:2px 3px;font-weight:500;">{a}</span>'
                    for a in terdeteksi)
                st.markdown(chips, unsafe_allow_html=True)
            else:
                st.caption("Tidak ada aspek spesifik yang terdeteksi.")

            # Catatan rekomendasi bila negatif
            if menang == "Negative" and terdeteksi:
                st.markdown(
                    f'<div style="background:#FEF3C7;border:1px solid #FDE68A;'
                    f'border-left:3px solid #D97706;border-radius:6px;'
                    f'padding:8px 12px;margin-top:8px;color:#92400E;'
                    f'font-size:0.83rem;">Ulasan negatif menyinggung aspek '
                    f'<b>{", ".join(terdeteksi)}</b> — aspek tersebut perlu '
                    f'menjadi perhatian UMKM.</div>',
                    unsafe_allow_html=True)

    # ── Detail teknis ──
    if hasil is not None:
        with st.expander("Detail Teknis"):
            st.markdown(f"**Model digunakan:** {model_pil}")
            st.markdown("**Teks asli:**")
            st.info(teks)
            st.markdown("**Teks setelah preprocessing:**")
            st.code(hasil["diproses"] if hasil["diproses"]
                    else "(kosong setelah preprocessing)", language=None)
            st.markdown(
                f"**Dimensi vektor TF-IDF:** {hasil['vec'].shape[1]} fitur")
            arr = hasil["vec"].toarray()[0]
            fitur = hasil["vectorizer"].get_feature_names_out()
            urut = np.argsort(arr)[::-1]
            top = [(fitur[i], round(float(arr[i]), 4))
                   for i in urut if arr[i] > 0][:10]
            st.markdown("**Top 10 kata paling berpengaruh (bobot TF-IDF):**")
            if top:
                st.dataframe(
                    pd.DataFrame(top, columns=["Kata / Fitur", "Bobot TF-IDF"]),
                    use_container_width=True, hide_index=True)
            else:
                st.caption("Tidak ada fitur yang dikenali dari ulasan ini.")


# =============================================================================
# HALAMAN 5 — PERBANDINGAN MODEL
# =============================================================================
def page_perbandingan():
    judul_halaman("Perbandingan Model Klasifikasi",
                  "Evaluasi performa tiga model ML pada dataset PRDECT-ID")
    st.markdown("<div style='height:0.6rem;'></div>", unsafe_allow_html=True)

    ev = load_evaluation()
    if not ev:
        st.info(
            "Hasil evaluasi belum tersedia. Jalankan terlebih dahulu:\n\n"
            "`python train.py`"
        )
        return

    MODEL_URUT   = ["Naive Bayes", "SVM", "Random Forest"]
    WARNA_MODEL  = {"Naive Bayes": "#6B7280", "SVM": "#B91C1C", "Random Forest": "#15803D"}
    best = ev.get("best_model", "SVM")
    ds   = ev.get("dataset_info", {})
    best_ev = ev.get(best, {})

    # ── Kartu model terbaik ─────────────────────────────────────────────────
    st.markdown(
        f'<div style="background:#FEF2F2;border:2px solid #FECACA;'
        f'border-left:5px solid #B91C1C;border-radius:10px;'
        f'padding:16px 20px;margin-bottom:1.4rem;display:flex;'
        f'align-items:center;gap:24px;">'
        f'<div>'
        f'<div style="font-size:0.75rem;color:#6B7280;text-transform:uppercase;'
        f'letter-spacing:.06em;font-weight:600;margin-bottom:4px;">Model Terbaik</div>'
        f'<div style="font-size:1.5rem;font-weight:800;color:#B91C1C;">{best}</div>'
        f'<div style="font-size:0.82rem;color:#6B7280;margin-top:4px;">'
        f'F1-Score Weighted tertinggi &middot; Cross-Validation stabil</div>'
        f'</div>'
        f'<div style="margin-left:auto;display:flex;gap:28px;text-align:center;">'
        f'<div><div style="font-size:1.3rem;font-weight:700;color:#B91C1C;">'
        f'{best_ev.get("accuracy",0)*100:.1f}%</div>'
        f'<div style="font-size:0.72rem;color:#6B7280;">Accuracy</div></div>'
        f'<div><div style="font-size:1.3rem;font-weight:700;color:#B91C1C;">'
        f'{best_ev.get("f1_weighted",0)*100:.1f}%</div>'
        f'<div style="font-size:0.72rem;color:#6B7280;">F1-Score</div></div>'
        f'<div><div style="font-size:1.3rem;font-weight:700;color:#B91C1C;">'
        f'{best_ev.get("cv_f1_mean",0)*100:.1f}%'
        f'<span style="font-size:0.8rem;font-weight:400;">'
        f' &plusmn;{best_ev.get("cv_f1_std",0)*100:.1f}</span></div>'
        f'<div style="font-size:0.72rem;color:#6B7280;">CV F1 (5-fold)</div></div>'
        f'</div></div>',
        unsafe_allow_html=True,
    )

    # ── Tabel perbandingan ──────────────────────────────────────────────────
    judul_seksi("Tabel Perbandingan Metrik")
    header_html = (
        f'<tr style="background:#F9FAFB;font-size:0.72rem;color:#6B7280;'
        f'text-transform:uppercase;letter-spacing:.05em;">'
        f'<th style="padding:10px 14px;text-align:left;">Model</th>'
        f'<th style="padding:10px 14px;text-align:center;">Accuracy</th>'
        f'<th style="padding:10px 14px;text-align:center;">F1 Weighted</th>'
        f'<th style="padding:10px 14px;text-align:center;">F1 Positif</th>'
        f'<th style="padding:10px 14px;text-align:center;">F1 Negatif</th>'
        f'<th style="padding:10px 14px;text-align:center;">CV F1 Mean</th>'
        f'<th style="padding:10px 14px;text-align:center;">CV Std</th>'
        f'</tr>'
    )
    baris_tabel = ""
    for m in MODEL_URUT:
        d = ev.get(m, {})
        is_best = (m == best)
        bg = "#FEF2F2" if is_best else "#FFFFFF"
        fw = "700" if is_best else "400"
        badge = (
            f' <span style="background:#B91C1C;color:#fff;font-size:0.65rem;'
            f'padding:1px 7px;border-radius:99px;font-weight:700;'
            f'vertical-align:middle;">Terbaik</span>' if is_best else ""
        )
        baris_tabel += (
            f'<tr style="background:{bg};border-top:1px solid #E5E7EB;">'
            f'<td style="padding:10px 14px;font-weight:{fw};color:#111827;">{m}{badge}</td>'
            f'<td style="padding:10px 14px;text-align:center;font-weight:{fw};">'
            f'{d.get("accuracy",0)*100:.2f}%</td>'
            f'<td style="padding:10px 14px;text-align:center;font-weight:{fw};">'
            f'{d.get("f1_weighted",0)*100:.2f}%</td>'
            f'<td style="padding:10px 14px;text-align:center;">'
            f'{d.get("f1_positive",0)*100:.2f}%</td>'
            f'<td style="padding:10px 14px;text-align:center;">'
            f'{d.get("f1_negative",0)*100:.2f}%</td>'
            f'<td style="padding:10px 14px;text-align:center;font-weight:{fw};">'
            f'{d.get("cv_f1_mean",0)*100:.2f}%</td>'
            f'<td style="padding:10px 14px;text-align:center;color:#6B7280;">'
            f'&plusmn;{d.get("cv_f1_std",0)*100:.2f}%</td>'
            f'</tr>'
        )
    st.markdown(
        f'<div style="overflow-x:auto;margin-bottom:1.4rem;">'
        f'<table style="width:100%;border-collapse:collapse;background:#fff;'
        f'border:1px solid #E5E7EB;border-radius:10px;overflow:hidden;font-size:0.875rem;">'
        f'{header_html}{baris_tabel}</table></div>',
        unsafe_allow_html=True,
    )

    # ── Bar chart ───────────────────────────────────────────────────────────
    judul_seksi("Visualisasi F1-Score per Model")
    metrik_label = ["F1 Weighted", "F1 Positif", "F1 Negatif", "CV F1 Mean"]
    metrik_key   = ["f1_weighted", "f1_positive", "f1_negative", "cv_f1_mean"]
    fig = go.Figure()
    for m in MODEL_URUT:
        d     = ev.get(m, {})
        nilai = [d.get(k, 0) * 100 for k in metrik_key]
        fig.add_trace(go.Bar(
            name=m + (" ★" if m == best else ""),
            x=metrik_label,
            y=nilai,
            marker_color=WARNA_MODEL[m],
            opacity=1.0 if m == best else 0.5,
            text=[f"{v:.1f}%" for v in nilai],
            textposition="outside",
            textfont=dict(size=11, color="#111827"),
        ))
    fig.update_layout(
        barmode="group",
        paper_bgcolor="#FFFFFF",
        plot_bgcolor="#FFFFFF",
        font=dict(family="Inter, sans-serif", color="#111827", size=12),
        yaxis=dict(title="Persentase (%)", range=[85, 100],
                   gridcolor="#E5E7EB", ticksuffix="%"),
        xaxis=dict(title="Metrik Evaluasi"),
        legend=dict(orientation="h", yanchor="bottom", y=1.02,
                    xanchor="right", x=1, bgcolor="rgba(0,0,0,0)"),
        margin=dict(t=40, b=40, l=40, r=20),
        height=400,
    )
    fig.update_xaxes(showgrid=False)
    st.plotly_chart(fig, use_container_width=True)

    # ── Confusion Matrix ────────────────────────────────────────────────────
    judul_seksi("Confusion Matrix per Model")
    st.caption("Baris = kelas aktual · Kolom = kelas prediksi · "
               "Hijau = prediksi benar · Merah = prediksi salah")
    cols = st.columns(3, gap="medium")
    for idx, m in enumerate(MODEL_URUT):
        d  = ev.get(m, {})
        cm = d.get("confusion_matrix", [[0, 0], [0, 0]])
        tn, fp = cm[0][0], cm[0][1]
        fn, tp = cm[1][0], cm[1][1]
        total  = tn + fp + fn + tp
        is_best = (m == best)
        bd = "#B91C1C" if is_best else "#E5E7EB"
        badge_html = (
            f' <span style="background:#B91C1C;color:#fff;font-size:0.62rem;'
            f'padding:1px 7px;border-radius:99px;font-weight:700;'
            f'margin-left:6px;">Terbaik</span>' if is_best else ""
        )
        with cols[idx]:
            st.markdown(
                f'<div style="border:1.5px solid {bd};border-radius:10px;'
                f'padding:16px;background:#FFFFFF;margin-bottom:0.5rem;">'
                f'<div style="font-weight:700;font-size:0.9rem;color:#111827;'
                f'margin-bottom:12px;">{m}{badge_html}</div>'
                f'<table style="width:100%;border-collapse:separate;'
                f'border-spacing:4px;font-size:0.82rem;text-align:center;">'
                f'<tr>'
                f'<td style="padding:4px;color:#6B7280;font-size:0.68rem;"></td>'
                f'<td style="padding:4px;color:#6B7280;font-size:0.68rem;'
                f'font-weight:600;">Pred. Negatif</td>'
                f'<td style="padding:4px;color:#6B7280;font-size:0.68rem;'
                f'font-weight:600;">Pred. Positif</td>'
                f'</tr>'
                f'<tr>'
                f'<td style="padding:4px;color:#6B7280;font-size:0.68rem;'
                f'font-weight:600;text-align:left;">Aktual Negatif</td>'
                f'<td style="padding:12px;background:#DCFCE7;color:#15803D;'
                f'font-weight:700;border-radius:6px;font-size:1.1rem;">{tn}</td>'
                f'<td style="padding:12px;background:#FEE2E2;color:#B91C1C;'
                f'font-weight:700;border-radius:6px;font-size:1.1rem;">{fp}</td>'
                f'</tr>'
                f'<tr>'
                f'<td style="padding:4px;color:#6B7280;font-size:0.68rem;'
                f'font-weight:600;text-align:left;">Aktual Positif</td>'
                f'<td style="padding:12px;background:#FEE2E2;color:#B91C1C;'
                f'font-weight:700;border-radius:6px;font-size:1.1rem;">{fn}</td>'
                f'<td style="padding:12px;background:#DCFCE7;color:#15803D;'
                f'font-weight:700;border-radius:6px;font-size:1.1rem;">{tp}</td>'
                f'</tr>'
                f'</table>'
                f'<div style="margin-top:10px;font-size:0.75rem;color:#6B7280;'
                f'border-top:1px solid #E5E7EB;padding-top:8px;">'
                f'Akurasi: <strong>{(tn+tp)/total*100:.1f}%</strong>'
                f' &nbsp;&middot;&nbsp; '
                f'Salah klasifikasi: <strong>{fp+fn}</strong> sampel'
                f'</div></div>',
                unsafe_allow_html=True,
            )

    # ── Info dataset ────────────────────────────────────────────────────────
    st.markdown(
        f'<div style="margin-top:1rem;background:#F9FAFB;border:1px solid #E5E7EB;'
        f'border-radius:8px;padding:14px 18px;font-size:0.82rem;color:#6B7280;'
        f'display:flex;gap:32px;flex-wrap:wrap;">'
        f'<span>Dataset: <strong style="color:#111827;">PRDECT-ID</strong></span>'
        f'<span>Total sampel: <strong style="color:#111827;">'
        f'{ds.get("total_samples",5400):,}</strong></span>'
        f'<span>Training: <strong style="color:#111827;">'
        f'{ds.get("train_samples",4320):,}</strong></span>'
        f'<span>Testing: <strong style="color:#111827;">'
        f'{ds.get("test_samples",1080):,}</strong></span>'
        f'<span>Fitur TF-IDF: <strong style="color:#111827;">'
        f'{ds.get("n_features",0):,}</strong></span>'
        f'<span>Validasi: <strong style="color:#111827;">'
        f'5-Fold Stratified CV</strong></span>'
        f'</div>',
        unsafe_allow_html=True,
    )


# =============================================================================
# HALAMAN 6 — LAPORAN
# =============================================================================
def _buat_xlsx_laporan(df: pd.DataFrame) -> bytes:
    """Bangun file Excel laporan UMKM dengan styling openpyxl.

    Layout:
      Row 1 : Judul (merge A:terakhir), merah tua #7F1D1D, teks putih, size 13
      Row 2 : Header tabel, merah #B91C1C, teks putih bold, tinggi 22
      Row 3+: Data — alternating putih / #FEF2F2, border tipis #E5E7EB,
              kolom Status diwarnai per nilai, kolom Skor center bold.
      Freeze panes di A2 (judul tetap terlihat saat scroll).
    """
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Laporan UMKM", startrow=1)
        ws = writer.sheets["Laporan UMKM"]

        n_cols = len(df.columns)
        n_rows = len(df)
        last_col = get_column_letter(n_cols)
        cols = list(df.columns)
        status_idx = cols.index("Status") + 1 if "Status" in cols else None
        skor_idx   = cols.index("Skor") + 1 if "Skor" in cols else None

        thin   = Side(style="thin", color="E5E7EB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # ── Row 1 : Judul ──
        ws.merge_cells(f"A1:{last_col}1")
        judul = ws["A1"]
        judul.value = "LAPORAN PEMANTAUAN UMKM - SIMANTAP DISKOP BANYUMAS"
        judul.fill = PatternFill("solid", fgColor="7F1D1D")
        judul.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
        judul.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # ── Row 2 : Header tabel ──
        header_fill  = PatternFill("solid", fgColor="B91C1C")
        header_font  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center",
                                 wrap_text=True)
        ws.row_dimensions[2].height = 22
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=2, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = border

        # ── Row 3+ : Data ──
        status_style = {
            "Kritis":          ("FEE2E2", "991B1B", True),
            "Perlu Perhatian": ("FFFBEB", "92400E", False),
            "Pantau":          ("F3F4F6", "374151", False),
            "Baik":            ("DCFCE7", "166534", False),
        }
        white_fill = PatternFill("solid", fgColor="FFFFFF")
        alt_fill   = PatternFill("solid", fgColor="FEF2F2")
        body_font  = Font(name="Calibri", size=11, color="111827")
        skor_font  = Font(name="Calibri", size=11, bold=True, color="111827")
        left_align   = Alignment(horizontal="left", vertical="center",
                                 wrap_text=True)
        center_align = Alignment(horizontal="center", vertical="center")

        for r in range(n_rows):
            excel_row = 3 + r
            row_fill = alt_fill if (r % 2 == 1) else white_fill
            for c in range(1, n_cols + 1):
                cell = ws.cell(row=excel_row, column=c)
                cell.border = border
                cell.fill = row_fill
                cell.font = body_font
                cell.alignment = left_align
                if c == status_idx:
                    style = status_style.get(str(cell.value))
                    if style:
                        bg, fg, bold = style
                        cell.fill = PatternFill("solid", fgColor=bg)
                        cell.font = Font(name="Calibri", size=11,
                                         bold=bold, color=fg)
                        cell.alignment = center_align
                elif c == skor_idx:
                    cell.font = skor_font
                    cell.alignment = center_align

        # ── Auto-fit lebar kolom (min 10, max 40) ──
        for c, col_name in enumerate(cols, start=1):
            max_len = len(str(col_name))
            for val in df.iloc[:, c - 1].astype(str):
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[get_column_letter(c)].width = max(
                10, min(40, max_len + 2)
            )

        # ── Freeze pane (judul tetap terlihat) ──
        ws.freeze_panes = "A2"

    return buf.getvalue()


def page_laporan():
    data = butuh_umkm_data()
    umkm = data["umkm"]
    now  = datetime.now()

    judul_halaman("Laporan Pemantauan UMKM",
                  f"Rekap {BULAN_ID[now.month]} {now.year}"
                  f" · {len(umkm)} UMKM terpantau")
    st.markdown("<div style='height:0.6rem;'></div>", unsafe_allow_html=True)

    jml = {s: sum(1 for u in umkm if u["status"] == s)
           for s in ["Kritis", "Perlu Perhatian", "Pantau", "Baik"]}

    lap_kiri, lap_kanan = st.columns([3, 2], gap="large")

    with lap_kiri:
        judul_seksi("Ringkasan per Status")
        c1, c2, c3, c4 = st.columns(4)
        for col, status in zip([c1, c2, c3, c4],
                               ["Kritis", "Perlu Perhatian", "Pantau", "Baik"]):
            s = STATUS_STYLE[status]
            with col:
                st.markdown(
                    f'<div style="background:{s["bg"]};border:1px solid {s["bd"]};'
                    f'border-radius:10px;padding:14px 10px;text-align:center;">'
                    f'<div style="font-size:2rem;font-weight:800;color:{s["fg"]};">'
                    f'{jml[status]}</div>'
                    f'<div style="font-size:0.75rem;color:{s["fg"]};font-weight:600;">'
                    f'{status}</div></div>',
                    unsafe_allow_html=True,
                )

        st.markdown("<div style='height:1rem;'></div>", unsafe_allow_html=True)
        judul_seksi("Distribusi Aspek Bermasalah")
        umkm_masalah = [u for u in umkm
                        if u["status"] in ["Kritis", "Perlu Perhatian"]]
        aspek_count = Counter()
        for u in umkm_masalah:
            for aspek, info in u.get("aspek", {}).items():
                # UMKM dihitung bermasalah di aspek X hanya jika aspek X
                # memang memiliki keluhan (bukan sekadar terdeteksi disebut).
                if isinstance(info, dict) and info.get("keluhan_count", 0) > 0:
                    aspek_count[aspek] += 1
        if aspek_count and umkm_masalah:
            urut = sorted(aspek_count.items(), key=lambda x: x[1], reverse=True)
            fig = go.Figure(go.Bar(
                x=[v / len(umkm_masalah) * 100 for _, v in urut],
                y=[a for a, _ in urut],
                orientation="h",
                marker_color=MERAH,
                text=[f"{v/len(umkm_masalah)*100:.1f}%" for _, v in urut],
                textposition="outside",
            ))
            fig.update_layout(
                paper_bgcolor="white", plot_bgcolor="white",
                font=dict(family="Inter, sans-serif", color=TEKS),
                xaxis=dict(title="% UMKM Bermasalah", range=[0, 115],
                           gridcolor=BORDER),
                yaxis=dict(title=""),
                margin=dict(t=10, b=30, l=80, r=60),
                height=220, template="plotly_white",
            )
            st.plotly_chart(fig, use_container_width=True)

    with lap_kanan:
        judul_seksi("Ringkasan per Kategori")
        kategori = sorted({u["kategori"] for u in umkm})
        rows = ""
        for kat in kategori:
            grup = [u for u in umkm if u["kategori"] == kat]
            kr = sum(1 for u in grup if u["status"] == "Kritis")
            pp = sum(1 for u in grup if u["status"] == "Perlu Perhatian")
            pt = sum(1 for u in grup if u["status"] == "Pantau")
            bk = sum(1 for u in grup if u["status"] == "Baik")
            rows += (
                f'<tr>'
                f'<td style="padding:8px 10px;border-bottom:1px solid {BORDER};">'
                f'{kat}</td>'
                f'<td style="padding:8px 10px;border-bottom:1px solid {BORDER};'
                f'text-align:center;font-weight:600;">{len(grup)}</td>'
                f'<td style="padding:8px 10px;border-bottom:1px solid {BORDER};'
                f'text-align:center;color:{MERAH};">{kr}</td>'
                f'<td style="padding:8px 10px;border-bottom:1px solid {BORDER};'
                f'text-align:center;color:#B45309;">{pp}</td>'
                f'<td style="padding:8px 10px;border-bottom:1px solid {BORDER};'
                f'text-align:center;color:{TEKS2};">{pt}</td>'
                f'<td style="padding:8px 10px;border-bottom:1px solid {BORDER};'
                f'text-align:center;color:{HIJAU};">{bk}</td>'
                f'</tr>'
            )
        st.markdown(
            f'<table style="width:100%;border-collapse:collapse;background:#FFFFFF;'
            f'border:1px solid {BORDER};border-radius:8px;overflow:hidden;'
            f'font-size:0.85rem;">'
            f'<tr style="background:#F9FAFB;color:{TEKS2};font-size:0.74rem;">'
            f'<th style="padding:8px 10px;text-align:left;">KATEGORI</th>'
            f'<th style="padding:8px 10px;">TOTAL</th>'
            f'<th style="padding:8px 10px;">KRITIS</th>'
            f'<th style="padding:8px 10px;">PERLU PERHATIAN</th>'
            f'<th style="padding:8px 10px;">PANTAU</th>'
            f'<th style="padding:8px 10px;">BAIK</th>'
            f'</tr>{rows}</table>',
            unsafe_allow_html=True,
        )

    st.markdown(
        f'<hr style="border:none;border-top:1px solid {BORDER};margin:24px 0;">',
        unsafe_allow_html=True,
    )

    judul_seksi("Tabel Lengkap UMKM")
    baris = []
    for u in umkm:
        rek = u["rekomendasi"]
        baris.append({
            "ID": u["id"],
            "Nama": u["nama"],
            "Kategori": u["kategori"],
            "Lokasi": u["lokasi"],
            "Skor": u["skor"],
            "Status": u["status"],
            "Masalah Utama": u["masalah_utama"],
            "Rekomendasi 1": rek[0] if len(rek) > 0 else "",
            "Rekomendasi 2": rek[1] if len(rek) > 1 else "",
        })
    df_laporan = pd.DataFrame(baris)
    st.dataframe(df_laporan, use_container_width=True, hide_index=True, height=380)

    csv_data = df_laporan.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
    xlsx_data = _buat_xlsx_laporan(df_laporan)

    dl_csv, dl_xlsx = st.columns(2)
    with dl_csv:
        st.download_button(
            label="⬇ Unduh Laporan CSV",
            data=csv_data,
            file_name=f"laporan_umkm_banyumas_{now.strftime('%Y%m')}.csv",
            mime="text/csv",
        )
    with dl_xlsx:
        st.download_button(
            label="⬇ Unduh Excel",
            data=xlsx_data,
            file_name=f"laporan_umkm_banyumas_{now.strftime('%Y%m')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


# =============================================================================
# ALUR UTAMA
# =============================================================================
if not st.session_state.login_ok:
    render_login()
    st.stop()

render_sidebar()

hal = st.session_state.halaman
if hal == "overview":
    page_overview()
elif hal == "daftar":
    page_daftar()
elif hal == "detail":
    page_detail()
elif hal == "analisis":
    page_analisis()
elif hal == "perbandingan":
    page_perbandingan()
elif hal == "laporan":
    page_laporan()
else:
    st.session_state.halaman = "overview"
    st.rerun()
